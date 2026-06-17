import re
from collections import Counter
from pathlib import Path

import openpyxl
import pandas as pd


BASE = Path(
    r"C:\Users\25124968R\Documents\Codex\2026-06-11\files-mentioned-by-the-user-all\outputs"
)
INPUT_CSV = BASE / "CUHK_OCT_OD_Quantize_figure1_values_all_rows.csv"
MATCH_XLSX = BASE / "CUHK_Record_260612_DOB_Age_cleaned.xlsx"
OUTPUT_CSV = BASE / "CUHK_OCT_OD_Quantize_figure1_values_all_rows_with_ID.csv"


def norm_name(value):
    return re.sub(r"\s+", " ", str(value or "").strip()).casefold()


def clean_id(value):
    text = str(value or "").strip()
    return re.sub(r"\.0$", "", text)


def clean_dob(value):
    text = re.sub(r"\.0$", "", str(value or "").strip())
    return re.sub(r"\D", "", text)[:8]


def sex_to_code(value):
    text = str(value or "").strip().upper()
    if text in {"F", "FEMALE", "1"}:
        return "1"
    if text in {"M", "MALE", "2"}:
        return "2"
    return ""


def parse_subject(subject):
    match = re.match(r"^(.*)_([0-9]{8})_([MF])$", str(subject).strip(), flags=re.I)
    if not match:
        return "", "", ""
    return match.group(1), match.group(2), sex_to_code(match.group(3))


wb = openpyxl.load_workbook(MATCH_XLSX, read_only=True, data_only=True)
ws = wb.active
headers = [ws.cell(2, col).value for col in range(1, ws.max_column + 1)]
name_col = headers.index("Name of Subject") + 1
id_col = headers.index("ID") + 1
dob_col = headers.index("DOB") + 1
sex_col = headers.index("Sex") + 1

excel_rows = []
for row in range(3, ws.max_row + 1):
    name = ws.cell(row, name_col).value
    identifier = ws.cell(row, id_col).value
    if name in (None, "") or identifier in (None, ""):
        continue
    excel_rows.append(
        {
            "name": str(name).strip(),
            "name_key": norm_name(name),
            "dob_sex_key": f"{clean_dob(ws.cell(row, dob_col).value)}_{sex_to_code(ws.cell(row, sex_col).value)}",
            "ID": clean_id(identifier),
        }
    )

name_counts = Counter(item["name_key"] for item in excel_rows)
dob_sex_counts = Counter(item["dob_sex_key"] for item in excel_rows)
id_by_name = {
    item["name_key"]: item["ID"] for item in excel_rows if name_counts[item["name_key"]] == 1
}
id_by_dob_sex = {
    item["dob_sex_key"]: item["ID"]
    for item in excel_rows
    if dob_sex_counts[item["dob_sex_key"]] == 1
}

df = pd.read_csv(INPUT_CSV, dtype=str, encoding="utf-8-sig")
df = df[df["source_subject"].ne("KU LOK KIU_20200420_F")].copy()

ids = []
methods = []
for _, item in df.iterrows():
    name_key = norm_name(item.get("Name", ""))
    _, dob, sex = parse_subject(item.get("source_subject", ""))
    dob_sex_key = f"{dob}_{sex}" if dob and sex else ""

    if name_key in id_by_name:
        ids.append(id_by_name[name_key])
        methods.append("name")
    elif dob_sex_key in id_by_dob_sex:
        ids.append(id_by_dob_sex[dob_sex_key])
        methods.append("dob_sex_unique_fallback")
    elif dob_sex_key and dob_sex_counts.get(dob_sex_key, 0) > 1:
        ids.append("")
        methods.append("duplicate_dob_sex_no_id")
    else:
        ids.append("")
        methods.append("unmatched_no_id")

df.insert(df.columns.get_loc("Name") + 1, "ID", ids)
df.insert(df.columns.get_loc("ID") + 1, "ID_match_method", methods)

df.to_csv(OUTPUT_CSV, index=False, encoding="utf-8-sig")

print(
    {
        "output": str(OUTPUT_CSV),
        "rows": int(len(df)),
        "columns": int(len(df.columns)),
        "removed_subject": "KU LOK KIU_20200420_F",
        "id_filled": int(pd.Series(ids).replace("", pd.NA).notna().sum()),
        "id_blank": int(pd.Series(ids).replace("", pd.NA).isna().sum()),
        "method_counts": dict(Counter(methods)),
    }
)
